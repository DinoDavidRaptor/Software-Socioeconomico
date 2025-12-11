"""
Módulo de exportación a Excel (XLSX).
Autor: DINOS Tech
Versión: 0.2.0
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List


class ExportadorExcel:
    """Clase para exportar múltiples estudios socioeconómicos a formato Excel."""
    
    def __init__(self, config_empresa: Dict):
        """
        Inicializa el exportador con la configuración de la empresa.
        
        Args:
            config_empresa: Diccionario con datos de la empresa.
        """
        self.config = config_empresa
    
    def _aplicar_estilos_encabezado(self, ws, row: int):
        """
        Aplica estilos al encabezado de la tabla.
        
        Args:
            ws: Hoja de trabajo.
            row: Número de fila del encabezado.
        """
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    def _aplicar_bordes(self, ws, start_row: int, end_row: int, start_col: int, end_col: int):
        """
        Aplica bordes a un rango de celdas.
        
        Args:
            ws: Hoja de trabajo.
            start_row, end_row: Rango de filas.
            start_col, end_col: Rango de columnas.
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, 
                               min_col=start_col, max_col=end_col):
            for cell in row:
                cell.border = thin_border
    
    def _colorear_riesgo(self, ws, cell, valor_riesgo: float):
        """
        Aplica color a una celda según el nivel de riesgo.
        
        Args:
            ws: Hoja de trabajo.
            cell: Celda a colorear.
            valor_riesgo: Valor del riesgo (1-5).
        """
        if valor_riesgo <= 1.5:
            fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Verde
        elif valor_riesgo <= 2.5:
            fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Amarillo
        elif valor_riesgo <= 3.5:
            fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # Naranja claro
        elif valor_riesgo <= 4.5:
            fill = PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid")  # Naranja
        else:
            fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Rojo
        
        cell.fill = fill
    
    def exportar(self, estudios_datos: List[Dict], ruta_salida: str) -> bool:
        """
        Exporta múltiples estudios a un archivo Excel con tabla comparativa.
        
        Args:
            estudios_datos: Lista de diccionarios con los datos de los estudios.
            ruta_salida: Ruta completa del archivo XLSX de salida.
            
        Returns:
            True si se exportó correctamente, False en caso contrario.
        """
        try:
            os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparativa de Estudios"
            
            # Encabezado de la empresa
            ws['A1'] = self.config.get("nombre", "")
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A1:AG1')
            
            ws['A2'] = f"Reporte Comparativo de Estudios Socioeconómicos - {datetime.now().strftime('%d/%m/%Y')}"
            ws['A2'].alignment = Alignment(horizontal="center")
            ws.merge_cells('A2:AG2')
            
            # Encabezados de columnas (fila 4)
            headers = [
                "Nombre Completo",
                "Edad",
                "Estado Civil",
                "Escolaridad",
                "Teléfono",
                "Email",
                "Empresa Actual",
                "Puesto",
                "Sueldo Mensual",
                "Total Gastos",
                "Balance",
                "% Gasto/Ingreso",
                "Núm. Hijos",
                "Núm. Dependientes",
                "Personas en Hogar",
                "Estado de Salud",
                "Enfermedades Crónicas",
                "Tipo Vivienda",
                "Tenencia Vivienda",
                "Riesgo Financiero",
                "Just. Financiero",
                "Riesgo Familiar",
                "Just. Familiar",
                "Riesgo Vivienda",
                "Just. Vivienda",
                "Riesgo Laboral",
                "Just. Laboral",
                "Riesgo Salud",
                "Just. Salud",
                "Riesgo Estilo Vida",
                "Just. Estilo Vida",
                "Riesgo Global",
                "Interpretación"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
            
            self._aplicar_estilos_encabezado(ws, 4)
            
            # Datos de estudios
            row_num = 5
            
            for estudio in estudios_datos:
                from src.logic.calculador_riesgos import CalculadorRiesgos
                
                dp = estudio.get("datos_personales", {})
                fin = estudio.get("situacion_financiera", {})
                fam = estudio.get("informacion_familiar", {})
                salud = estudio.get("salud_intereses", {})
                viv = estudio.get("vivienda", {})
                empleo = estudio.get("empleo_actual", {})
                
                # Calcular riesgos con justificaciones
                calc = CalculadorRiesgos(estudio)
                resultados = calc.calcular_todos_riesgos()
                
                # Calcular porcentaje gasto/ingreso
                sueldo = fin.get("sueldo_mensual", 0)
                otros_ingresos = sum(ing.get("monto", 0) for ing in fin.get("otros_ingresos", []))
                ingreso_total = sueldo + otros_ingresos
                gastos_total = fin.get("gastos", {}).get("total", 0)
                porcentaje_gasto = (gastos_total / ingreso_total * 100) if ingreso_total > 0 else 0
                
                # Contar enfermedades crónicas
                enfermedades = salud.get("enfermedades_cronicas", [])
                enfermedades_texto = ", ".join([e.get("nombre", "") for e in enfermedades]) if enfermedades else "Ninguna"
                
                # Extraer justificaciones
                def get_justificaciones(key):
                    if key in resultados and resultados[key]["justificaciones"]:
                        return " | ".join(resultados[key]["justificaciones"])
                    return ""
                
                # Escribir datos
                datos_fila = [
                    dp.get("nombre_completo", "N/A"),
                    dp.get("edad", 0),
                    dp.get("estado_civil", "N/A"),
                    dp.get("escolaridad", "N/A"),
                    dp.get("telefono", "N/A"),
                    dp.get("email", "N/A"),
                    empleo.get("empresa", fin.get("empresa_actual", "N/A")),
                    empleo.get("puesto", fin.get("puesto_actual", "N/A")),
                    fin.get("sueldo_mensual", 0),
                    gastos_total,
                    fin.get("balance", 0),
                    porcentaje_gasto,
                    fam.get("numero_hijos", 0),
                    fam.get("numero_dependientes_economicos", 0),
                    fam.get("personas_hogar", 0),
                    salud.get("estado_salud", "N/A"),
                    enfermedades_texto,
                    viv.get("tipo_vivienda", "N/A"),
                    viv.get("tenencia", "N/A"),
                    resultados.get("financiero", {}).get("puntaje", 0),
                    get_justificaciones("financiero"),
                    resultados.get("familiar", {}).get("puntaje", 0),
                    get_justificaciones("familiar"),
                    resultados.get("vivienda", {}).get("puntaje", 0),
                    get_justificaciones("vivienda"),
                    resultados.get("laboral", {}).get("puntaje", 0),
                    get_justificaciones("laboral"),
                    resultados.get("salud", {}).get("puntaje", 0),
                    get_justificaciones("salud"),
                    resultados.get("estilo_vida", {}).get("puntaje", 0),
                    get_justificaciones("estilo_vida"),
                    resultados.get("global", {}).get("puntaje", 0),
                    CalculadorRiesgos.obtener_interpretacion_riesgo(resultados.get("global", {}).get("puntaje", 0))
                ]
                
                for col, valor in enumerate(datos_fila, 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = valor
                    
                    # Alineación y formato
                    if col in [9, 10, 11]:  # Montos
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '"$"#,##0.00'
                    elif col == 12:  # Porcentaje
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '0.00"%"'
                    elif col in [20, 22, 24, 26, 28, 30, 32]:  # Columnas de puntajes de riesgo
                        cell.alignment = Alignment(horizontal="center")
                        # Aplicar color según el nivel de riesgo
                        self._colorear_riesgo(ws, cell, valor)
                    elif col in [21, 23, 25, 27, 29, 31]:  # Columnas de justificaciones
                        cell.alignment = Alignment(horizontal="left", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left")
                
                row_num += 1
            
            # Aplicar bordes
            self._aplicar_bordes(ws, 4, row_num - 1, 1, len(headers))
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 25,  # Nombre
                'B': 8,   # Edad
                'C': 15,  # Estado Civil
                'D': 20,  # Escolaridad
                'E': 15,  # Teléfono
                'F': 25,  # Email
                'G': 20,  # Empresa
                'H': 20,  # Puesto
                'I': 15,  # Sueldo
                'J': 15,  # Gastos
                'K': 15,  # Balance
                'L': 12,  # % Gasto
                'M': 10,  # Hijos
                'N': 12,  # Dependientes
                'O': 12,  # Personas Hogar
                'P': 15,  # Estado Salud
                'Q': 25,  # Enfermedades
                'R': 15,  # Tipo Vivienda
                'S': 15,  # Tenencia
                'T': 10,  # R. Financiero
                'U': 50,  # Just. Financiero
                'V': 10,  # R. Familiar
                'W': 50,  # Just. Familiar
                'X': 10,  # R. Vivienda
                'Y': 50,  # Just. Vivienda
                'Z': 10,  # R. Laboral
                'AA': 50, # Just. Laboral
                'AB': 10, # R. Salud
                'AC': 50, # Just. Salud
                'AD': 10, # R. Estilo Vida
                'AE': 50, # Just. Estilo Vida
                'AF': 12, # R. Global
                'AG': 20  # Interpretación
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Aumentar altura de filas para justificaciones
            for row in range(5, row_num):
                ws.row_dimensions[row].height = 60
            
            # Congelar paneles (fijar encabezados)
            ws.freeze_panes = 'A5'
            
            # Agregar leyenda de colores
            leyenda_row = row_num + 2
            ws.cell(row=leyenda_row, column=1).value = "Leyenda de Riesgos:"
            ws.cell(row=leyenda_row, column=1).font = Font(bold=True)
            
            leyenda_items = [
                ("Muy Bajo (1-1.5)", "C8E6C9"),
                ("Bajo (1.6-2.5)", "FFF9C4"),
                ("Medio (2.6-3.5)", "FFE0B2"),
                ("Alto (3.6-4.5)", "FFCCBC"),
                ("Muy Alto (4.6-5)", "FFCDD2")
            ]
            
            for idx, (texto, color) in enumerate(leyenda_items):
                cell = ws.cell(row=leyenda_row + idx + 1, column=1)
                cell.value = texto
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Guardar archivo
            wb.save(ruta_salida)
            
            return True
            
        except Exception as e:
            print(f"Error al exportar Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
